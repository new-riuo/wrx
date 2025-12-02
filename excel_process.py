import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import os
import json
from datetime import datetime

class ExcelProcessor:
    """报关单管理系统Excel处理类"""
    
    # 常量定义
    THIS_WORKBOOK_NAME = "当前工作簿.xlsx"
    CONFIG_FILE_NAME = "template_config.json"
    
    def __init__(self):
        """初始化Excel处理器"""
        self.this_workbook = None
        self.text_box1 = ""
        # 数据存储
        self.logistics_data = []  # 物流信息数据
        self.order_data = []  # 订单信息数据
        self.declaration_info = []  # 报关信息数据
        self.country_codes = []  # 国家代码数据
        self.currency_data = []  # 币种数据
        self.shop_company_data = []  # 店铺对应公司数据
        self.declaration_amount_rules = []  # 申报金额规则数据
        self.customs_data = []  # 报关单数据
        # 加载配置
        self.load_config()
        # 初始化默认数据
        self.init_default_data()
    
    def init_default_data(self):
        """初始化默认数据"""
        # 初始化报关单模板字段
        self.declaration_template_fields = [
            {"field_name": "提单号", "field_type": "文本", "required": True},
            {"field_name": "订单编号", "field_type": "文本", "required": True},
            {"field_name": "大箱号", "field_type": "文本", "required": False},
            {"field_name": "快件单号", "field_type": "文本", "required": False},
            {"field_name": "包裹号（运单号）", "field_type": "文本", "required": True},
            {"field_name": "目的国", "field_type": "文本", "required": True},
            {"field_name": "敏感品类别", "field_type": "文本", "required": False},
            {"field_name": "商品品名", "field_type": "文本", "required": True},
            {"field_name": "HS CODE", "field_type": "文本", "required": True},
            {"field_name": "规格型号", "field_type": "文本", "required": False},
            {"field_name": "包裹内单个SKC的商品数量", "field_type": "数值", "required": True},
            {"field_name": "申报单位", "field_type": "文本", "required": True},
            {"field_name": "商品申报单价", "field_type": "数值", "required": True},
            {"field_name": "申报币制", "field_type": "文本", "required": True},
            {"field_name": "商品总净重(KG)", "field_type": "数值", "required": True},
            {"field_name": "第一法定数量", "field_type": "数值", "required": False},
            {"field_name": "第一法定单位", "field_type": "文本", "required": False},
            {"field_name": "第二法定数量", "field_type": "数值", "required": False},
            {"field_name": "第二法定单位", "field_type": "文本", "required": False},
            {"field_name": "电商企业代码", "field_type": "文本", "required": True},
            {"field_name": "电商企业名称", "field_type": "文本", "required": True},
            {"field_name": "电商平台代码", "field_type": "文本", "required": True},
            {"field_name": "电商平台名称", "field_type": "文本", "required": True},
            {"field_name": "收款企业代码", "field_type": "文本", "required": True},
            {"field_name": "收款企业名称", "field_type": "文本", "required": True},
            {"field_name": "生产企业代码", "field_type": "文本", "required": True},
            {"field_name": "生产企业名称", "field_type": "文本", "required": True},
            {"field_name": "电商企业dxpId", "field_type": "文本", "required": True}
        ]
        
        # 初始化默认国家代码（只包含consignee_country和three_letter_code）
        self.country_codes = [
            {"consignee_country": "United States", "three_letter_code": "502"},
            {"consignee_country": "United Kingdom", "three_letter_code": "303"},
            {"consignee_country": "GB", "three_letter_code": "303"},
            {"consignee_country": "United Kingdom", "three_letter_code": "303"},
        ]
        
        # 初始化默认币种数据
        self.currency_data = [
            {"country_name": "美国", "currency_code": "USD"},
            {"country_name": "英国", "currency_code": "GBP"},
            {"country_name": "加拿大", "currency_code": "CAD"},
            {"country_name": "澳大利亚", "currency_code": "AUD"},
            {"country_name": "德国", "currency_code": "EUR"},
            {"country_name": "法国", "currency_code": "EUR"},
            {"country_name": "意大利", "currency_code": "EUR"},
            {"country_name": "西班牙", "currency_code": "EUR"},
            {"country_name": "日本", "currency_code": "JPY"},
            {"country_name": "韩国", "currency_code": "KRW"},
            {"country_name": "新加坡", "currency_code": "SGD"},
            {"country_name": "马来西亚", "currency_code": "MYR"},
            {"country_name": "印度", "currency_code": "INR"},
            {"country_name": "巴西", "currency_code": "BRL"},
            {"country_name": "俄罗斯", "currency_code": "RUB"},
            {"country_name": "南非", "currency_code": "ZAR"}
        ]
        
        # 初始化默认店铺对应公司数据
        self.shop_company_data = [
            {"shop_name": "店铺1", "company_name": "公司A"},
            {"shop_name": "店铺2", "company_name": "公司B"},
            {"shop_name": "店铺3", "company_name": "公司C"}
        ]
        
        # 初始化默认申报金额规则
        self.declaration_amount_rules = [
            {"country_name": "美国", "declaration_ratio": 1, "max_declaration_amount": 3},
            {"country_name": "英国", "declaration_ratio": 1, "max_declaration_amount": 135},
            {"country_name": "加拿大", "declaration_ratio": 1, "max_declaration_amount": 20},
            {"country_name": "澳大利亚", "declaration_ratio": 0.5, "max_declaration_amount": 1000},
            {"country_name": "德国", "declaration_ratio": 0.6, "max_declaration_amount": 150},
            {"country_name": "法国", "declaration_ratio": 0.6, "max_declaration_amount": 150},
            {"country_name": "意大利", "declaration_ratio": 0.6, "max_declaration_amount": 150},
            {"country_name": "西班牙", "declaration_ratio": 0.6, "max_declaration_amount": 150}
        ]
        
        # 加载国家代码数据
        self.load_country_codes()
        # 加载币种数据
        self.load_currency_data()
        # 加载店铺对应公司数据
        self.load_shop_company_data()
        # 加载申报金额规则数据
        self.load_declaration_amount_rules()
        # 加载默认报关信息数据
        self.load_default_declaration_data()
    
    def load_default_declaration_data(self):
        """从CSV文件加载默认报关信息数据"""
        csv_file_path = "e:/xsfx/王如鑫/报关信息导入.csv"
        if not os.path.exists(csv_file_path):
            print(f"默认报关信息文件不存在: {csv_file_path}")
            return
        
        try:
            # 清空现有数据
            self.declaration_info = []
            
            # 使用pandas读取CSV文件，支持各种格式和编码
            import pandas as pd
            
            # 读取CSV文件，跳过空行，使用第一行作为表头
            df = pd.read_csv(csv_file_path, encoding='utf-8', skip_blank_lines=True)
            
            # 打印表头信息，用于调试
            print(f"CSV文件表头: {list(df.columns)}")
            
            # 遍历数据行
            for index, row in df.iterrows():
                # 过滤掉提单号为空的行
                if pd.isna(row.iloc[0]):
                    continue
                
                # 创建报关信息字典，只取前19列数据（根据CSV文件的实际结构）
                declaration_data = {
                    "提单号": str(row.iloc[0]) if not pd.isna(row.iloc[0]) else "",
                    "商品品名": str(row.iloc[1]) if not pd.isna(row.iloc[1]) else "",
                    "HS CODE": str(row.iloc[2]) if not pd.isna(row.iloc[2]) else "",
                    "规格型号": str(row.iloc[3]) if not pd.isna(row.iloc[3]) else "",
                    "包裹内单个SKC的商品数量": int(row.iloc[4]) if not pd.isna(row.iloc[4]) else 0,
                    "申报单位": str(row.iloc[5]) if not pd.isna(row.iloc[5]) else "",
                    "申报币制": str(row.iloc[6]) if not pd.isna(row.iloc[6]) else "",
                    "商品总净重(KG)": float(row.iloc[7]) if not pd.isna(row.iloc[7]) else 0.0,
                    "第一法定数量": float(row.iloc[8]) if not pd.isna(row.iloc[8]) else 0.0,
                    "第一法定单位": str(row.iloc[9]) if not pd.isna(row.iloc[9]) else "",
                    "电商企业代码": str(row.iloc[10]) if not pd.isna(row.iloc[10]) else "",
                    "电商企业名称": str(row.iloc[11]) if not pd.isna(row.iloc[11]) else "",
                    "电商平台代码": str(row.iloc[12]) if not pd.isna(row.iloc[12]) else "",
                    "电商平台名称": str(row.iloc[13]) if not pd.isna(row.iloc[13]) else "",
                    "收款企业代码": str(row.iloc[14]) if not pd.isna(row.iloc[14]) else "",
                    "收款企业名称": str(row.iloc[15]) if not pd.isna(row.iloc[15]) else "",
                    "生产企业代码": str(row.iloc[16]) if not pd.isna(row.iloc[16]) else "",
                    "生产企业名称": str(row.iloc[17]) if not pd.isna(row.iloc[17]) else "",
                    "电商企业dxpId": str(row.iloc[18]) if not pd.isna(row.iloc[18]) else ""
                }
                
                # 添加到报关信息列表
                self.declaration_info.append(declaration_data)
            
            print(f"成功加载 {len(self.declaration_info)} 条默认报关信息")
        except Exception as e:
            print(f"加载默认报关信息时发生错误: {e}")
            import traceback
            traceback.print_exc()
    
    def load_config(self):
        """从文件加载配置"""
        # 默认配置
        default_template_config = {
            "WA": {"sheet_name": "翼速特瑞福9610模板", "bm_range": "A3:AC3"},
            "假发3店": {"sheet_name": "翼速鑫瑞祥和9610模板", "bm_range": "A4:AC4"},
            "WA海外托管": {"sheet_name": "翼速鑫瑞祥和9610模板", "bm_range": "A4:AC4"},
            "假发5店": {"sheet_name": "翼速鑫瑞祥和9610模板", "bm_range": "A4:AC4"},
            "假发1店": {"sheet_name": "翼速碧欧丝9610模板", "bm_range": "A5:AC5"},
            "Roselover 2店": {"sheet_name": "翼速碧欧丝9610模板", "bm_range": "A5:AC5"},
            "Top Unique Hair": {"sheet_name": "翼速新启航9610模板", "bm_range": "A2:AC2"},
            "BS": {"sheet_name": "翼速新启航9610模板", "bm_range": "A2:AC2"},
            "AB": {"sheet_name": "翼速新启航9610模板", "bm_range": "A2:AC2"}
        }
        
        default_template_file_map = {
            "翼速特瑞福9610模板": "翼速特瑞福9610模板",
            "翼速鑫瑞祥和9610模板": "翼速鑫瑞祥和9610模板",
            "翼速新启航9610模板": "翼速新启航9610模板",
            "翼速碧欧丝9610模板": "翼速碧欧丝9610模板"
        }
        
        # 从文件加载配置
        if os.path.exists(self.CONFIG_FILE_NAME):
            try:
                with open(self.CONFIG_FILE_NAME, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.TEMPLATE_CONFIG = config.get("TEMPLATE_CONFIG", default_template_config)
                self.TEMPLATE_FILE_MAP = config.get("TEMPLATE_FILE_MAP", default_template_file_map)
            except Exception as e:
                print(f"加载配置文件失败，使用默认配置：{e}")
                self.TEMPLATE_CONFIG = default_template_config
                self.TEMPLATE_FILE_MAP = default_template_file_map
        else:
            # 使用默认配置
            self.TEMPLATE_CONFIG = default_template_config
            self.TEMPLATE_FILE_MAP = default_template_file_map
            # 保存默认配置到文件
            self.save_config()
    
    def save_config(self):
        """保存配置到文件"""
        try:
            config = {
                "TEMPLATE_CONFIG": self.TEMPLATE_CONFIG,
                "TEMPLATE_FILE_MAP": self.TEMPLATE_FILE_MAP
            }
            with open(self.CONFIG_FILE_NAME, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
            print("配置保存成功")
        except Exception as e:
            print(f"保存配置文件失败：{e}")
        
    def select_file(self, title="请选择文件"):
        """模拟Application.GetOpenFilename功能，选择Excel文件
        
        Args:
            title: 文件选择对话框标题
            
        Returns:
            str: 选中的文件路径，如果取消则返回空字符串
        """
        # 避免创建新的Tk()实例，而是使用现有实例或创建一个临时隐藏的Toplevel
        # 这样可以避免干扰主窗口的事件循环
        try:
            # 检查是否已经有root窗口存在
            if tk._default_root is not None:
                # 使用现有root窗口作为父窗口创建一个隐藏的Toplevel
                root = tk.Toplevel(tk._default_root)
                root.withdraw()  # 隐藏窗口
            else:
                # 如果没有root窗口，创建一个临时的
                root = tk.Tk()
                root.withdraw()  # 隐藏窗口
                root.after(0, root.destroy)  # 确保窗口会被销毁
            
            file_path = filedialog.askopenfilename(
                parent=root,
                title=title,
                filetypes=[("Excel文件", "*.xls;*.xlsx;*.xlsm"), ("所有文件", "*.*")]
            )
            
            return file_path
        except Exception as e:
            print(f"文件选择出错: {e}")
            # 作为最后的后备方案
            try:
                temp_root = tk.Tk()
                temp_root.withdraw()
                file_path = filedialog.askopenfilename(
                    title=title,
                    filetypes=[("Excel文件", "*.xls;*.xlsx;*.xlsm"), ("所有文件", "*.*")]
                )
                temp_root.destroy()
                return file_path
            except:
                return ""
    
    def import_logistics_file(self):
        """导入物流信息表 - 增强版"""
        try:
            # 选择物流信息表文件
            file_path = self.select_file(title="请选择物流信息表")
            
            if not file_path:
                return False  # 返回False表示未导入
            
            # 清空现有物流数据
            self.logistics_data = []
            
            # 使用pandas读取Excel文件，支持.xls和.xlsx格式
            try:
                import pandas as pd
                
                # 读取Excel文件
                df = pd.read_excel(file_path)
                
                # 获取表头并清理列名
                headers = df.columns.tolist()
                # 清理表头名称，去掉特殊字符和空格
                cleaned_headers = []
                for header in headers:
                    # 处理可能的Unnamed列
                    if isinstance(header, str) and 'Unnamed' in header:
                        cleaned_headers.append(f"列_{len(cleaned_headers)+1}")
                    else:
                        cleaned_headers.append(str(header).strip())
                
                # 读取数据行
                valid_rows = 0
                for index, row in df.iterrows():
                    # 跳过空行
                    if pd.isna(row.iloc[0]):
                        continue
                    
                    # 跳过可能的表头重复行
                    if index > 0 and str(row.iloc[0]).strip() == cleaned_headers[0]:
                        continue
                    
                    # 创建数据字典
                    row_data = {}
                    for i, header in enumerate(cleaned_headers):
                        # 处理单元格数据，确保所有值都是可序列化的类型
                        cell_value = row.iloc[i] if i < len(row) else None
                        # 处理NaN值
                        if pd.isna(cell_value):
                            row_data[header] = ""
                        # 处理日期时间类型
                        elif isinstance(cell_value, (pd.Timestamp, datetime)):
                            row_data[header] = cell_value.strftime("%Y-%m-%d %H:%M:%S")
                        # 处理数字类型，避免科学计数法
                        elif isinstance(cell_value, (int, float)):
                            # 对于整数，转换为整数类型
                            if cell_value == int(cell_value):
                                row_data[header] = int(cell_value)
                            else:
                                # 对于小数，保留原精度
                                row_data[header] = cell_value
                        else:
                            # 其他类型转换为字符串
                            row_data[header] = str(cell_value).strip()
                    
                    # 添加数据索引和标记
                    row_data["_index"] = index  # 添加原始索引
                    row_data["is_matched"] = True  # 默认标记为匹配
                    
                    self.logistics_data.append(row_data)
                    valid_rows += 1
                
                # 数据验证：检查是否真的有数据被导入
                if len(self.logistics_data) > 0:
                    return True  # 返回True表示导入成功
                else:
                    return False
                
            except Exception as e:
                print(f"读取Excel文件失败：{e}")
                return False  # 返回False表示导入失败
            
        except Exception as e:
            # 发生异常时清空数据，避免部分数据导入导致的问题
            self.logistics_data = []
            raise
    
    def process_logistics_data(self):
        """处理物流信息，抓取实际发货物流字段的"US"和红底的"DHL"并保存"""
        try:
            # 遍历物流数据，处理实际发货物流字段
            for row in self.logistics_data:
                # 检查是否有"实际发货物流"字段
                if "实际发货物流" in row:
                    logistics_value = row["实际发货物流"]
                    
                    # 初始化处理结果
                    row["处理结果"] = ""
                    
                    # 检查是否包含"US"
                    if "US" in str(logistics_value):
                        row["处理结果"] += "US "
                    
                    # 检查是否包含红底的"DHL"（这里简化处理，实际需要检查单元格格式）
                    if "DHL" in str(logistics_value):
                        row["处理结果"] += "DHL"
            
            print("物流信息处理完成")
        except Exception as e:
            print(f"处理物流信息时发生错误: {e}")
            raise
    
    def get_logistics_data(self):
        """获取物流数据
        
        Returns:
            list: 物流数据列表
        """
        return self.logistics_data
    
    def import_order_file(self):
        """导入订单信息表"""
        try:
            # 选择订单信息表文件
            file_path = self.select_file(title="请选择订单信息表")
            
            if not file_path:
                print("你没有选择文件")
                return
            
            # 清空现有订单数据
            self.order_data = []
            
            # 使用pandas读取Excel文件，支持.xls和.xlsx格式
            try:
                import pandas as pd
                
                # 读取Excel文件
                df = pd.read_excel(file_path)
                
                # 获取表头
                headers = df.columns.tolist()
                
                # 读取数据行
                for index, row in df.iterrows():
                    if pd.isna(row.iloc[0]):  # 跳过空行
                        continue
                    
                    # 创建数据字典
                    row_data = {}
                    for i, header in enumerate(headers):
                        row_data[header] = row.iloc[i]
                    
                    self.order_data.append(row_data)
                    
            except Exception as e:
                print(f"读取Excel文件失败：{e}")
                return
            
            print("订单信息表导入成功")
        except Exception as e:
            print(f"导入订单信息表时发生错误: {e}")
            raise
    
    def match_order_data(self):
        """根据物流信息匹配订单数据"""
        try:
            # 初始化所有订单数据的匹配标记为False
            for order_row in self.order_data:
                order_row["is_matched"] = False
            
            # 遍历物流数据，匹配订单数据
            for logistics_row in self.logistics_data:
                # 初始化匹配标记
                logistics_row["is_matched"] = False
                # 获取物流信息中的平台订单号
                if "平台订单号" in logistics_row:
                    platform_order_no = logistics_row["平台订单号"]
                    
                    # 查找匹配的订单数据
                    for order_row in self.order_data:
                        # 获取订单信息中的订单号（支持"订单号"和"order code"两种字段名）
                        order_no = order_row.get("订单号", "") or order_row.get("order code", "") or order_row.get("订单号/Order Code", "")
                        # 匹配条件：订单号/order code 与 平台订单号 相等
                        if order_no == platform_order_no:
                            # 将匹配的订单数据合并到物流数据中
                            for key, value in order_row.items():
                                if key not in logistics_row:
                                    logistics_row[key] = value
                            # 设置匹配标记为True
                            logistics_row["is_matched"] = True
                            order_row["is_matched"] = True
                            break
            
            print("订单数据匹配完成")
        except Exception as e:
            print(f"匹配订单数据时发生错误: {e}")
            raise
    
    def generate_declaration_data(self):
        """生成报关单数据
        
        Returns:
            list: 报关单数据列表
        """
        declaration_data = []
        total_orders = 0
        processed_orders = 0
        error_orders = 0
        
        try:
            # 检查必要的数据结构是否存在
            if not hasattr(self, 'order_data') or self.order_data is None:
                self.order_data = []
            
            if not hasattr(self, 'declaration_info'):
                self.declaration_info = []
            
            if not hasattr(self, 'shop_company_data'):
                self.shop_company_data = []
            
            if not hasattr(self, 'country_codes'):
                self.country_codes = []
            
            if not hasattr(self, 'declaration_amount_rules'):
                self.declaration_amount_rules = []
            
            # 确保有报关信息可用，如果没有则创建默认报关信息
            if not self.declaration_info:
                try:
                    # 创建默认报关信息
                    self.declaration_info = [{
                        "提单号": "",
                        "商品品名": "默认商品",
                        "HS CODE": "000000",
                        "规格型号": "",
                        "包裹内单个SKC的商品数量": 1,
                        "申报单位": "个",
                        "申报币制": "502",
                        "商品总净重(KG)": 0.3,
                        "第一法定数量": 0.3,
                        "第一法定单位": "035",
                        "电商企业代码": "",
                        "电商企业名称": "默认企业名称",
                        "电商平台代码": "",
                        "电商平台名称": "",
                        "收款企业代码": "",
                        "收款企业名称": "",
                        "生产企业代码": "",
                        "生产企业名称": "",
                        "电商企业dxpId": ""
                    }]
                except Exception as e:
                    print(f"创建默认报关信息时出错: {str(e)}")
            
            # 如果没有订单数据，创建一个默认订单用于测试
            if not self.order_data:
                try:
                    self.order_data = [{
                        "订单编号": "测试订单001",
                        "Order Code": "TEST001",
                        "目的国": "美国",
                        "物流跟踪号": "TEST-TRACKING-001",
                        "总销售金额": 100.0,
                        "AmountPaid": 100.0,
                        "店铺名称": "Top Unique Hair",
                        "Consignee Country": "United States"
                    }]
                except Exception as e:
                    print(f"创建默认测试订单时出错: {str(e)}")
            
            # 确定数据来源：优先使用匹配后的物流数据，否则使用订单数据
            data_source = []
            use_logistics_data = False
            
            if hasattr(self, 'logistics_data') and self.logistics_data:
                # 使用匹配后的物流数据
                data_source = self.logistics_data
                use_logistics_data = True
            else:
                # 没有物流数据，使用订单数据
                data_source = self.order_data
            
            total_orders = len(data_source)
            
            # 遍历数据来源，生成报关单数据
            for index, data_row in enumerate(data_source):
                try:
                    # 检查数据行是否为字典类型
                    if not isinstance(data_row, dict):
                        error_orders += 1
                        continue
                    
                    # 如果使用物流数据，检查是否已匹配到订单
                    if use_logistics_data:
                        if not data_row.get("is_matched", False):
                            error_orders += 1
                            continue
                    
                    # 获取订单编号，支持多种字段名
                    try:
                        order_no = data_row.get("订单号", "") or \
                                  data_row.get("order code", "") or \
                                  data_row.get("订单号/Order Code", "") or \
                                  data_row.get("订单编号", "") or \
                                  data_row.get("Order Code", "") or \
                                  data_row.get("平台订单号", "") or \
                                  f"默认订单号_{index+1}"
                    except Exception as e:
                        order_no = f"错误订单号_{index+1}"
                    
                    # 获取店铺名称
                    try:
                        shop_name = data_row.get("店铺名称", "默认店铺")
                    except Exception as e:
                        shop_name = "默认店铺"
                    
                    # 根据店铺名称查找对应的公司名称
                    try:
                        company_name = "默认企业名称"
                        for shop_company in self.shop_company_data:
                            if isinstance(shop_company, dict) and shop_company.get("shop_name") == shop_name:
                                company_name = shop_company.get("company_name", "默认企业名称")
                                break
                    except Exception as e:
                        company_name = "默认企业名称"
                    
                    # 根据公司名称查找对应的报关信息 - 重点修改：按照README.md要求
                    try:
                        company_declaration_info = None
                        # 优先使用与订单号所属店铺的公司匹配的报关信息
                        for declaration in self.declaration_info:
                            if isinstance(declaration, dict) and declaration.get("电商企业名称", "") == company_name:
                                company_declaration_info = declaration
                                break
                        
                        # 如果没有找到匹配的报关信息，使用第一条作为默认值
                        if not company_declaration_info and self.declaration_info:
                            company_declaration_info = self.declaration_info[0]
                        
                        # 如果仍然没有报关信息，创建一个默认的
                        if not company_declaration_info:
                            company_declaration_info = {
                                "提单号": "",
                                "商品品名": "默认商品",
                                "HS CODE": "000000",
                                "规格型号": "",
                                "包裹内单个SKC的商品数量": 1,
                                "申报单位": "个",
                                "申报币制": "502",
                                "商品总净重(KG)": 0.3,
                                "第一法定数量": 0.3,
                                "第一法定单位": "035",
                                "电商企业代码": "",
                                "电商企业名称": company_name,
                                "电商平台代码": "",
                                "电商平台名称": "",
                                "收款企业代码": "",
                                "收款企业名称": "",
                                "生产企业代码": "",
                                "生产企业名称": "",
                                "电商企业dxpId": ""
                            }
                    except Exception as e:
                        # 创建最小化的默认报关信息以保证程序继续运行
                        company_declaration_info = {
                            "提单号": "",
                            "商品品名": "默认商品",
                            "HS CODE": "000000",
                            "规格型号": "",
                            "包裹内单个SKC的商品数量": 1,
                            "申报单位": "个",
                            "申报币制": "502",
                            "商品总净重(KG)": 0.3,
                            "第一法定数量": 0.3,
                            "第一法定单位": "035",
                            "电商企业代码": "",
                            "电商企业名称": company_name,
                            "电商平台代码": "",
                            "电商平台名称": "",
                            "收款企业代码": "",
                            "收款企业名称": "",
                            "生产企业代码": "",
                            "生产企业名称": "",
                            "电商企业dxpId": ""
                        }
                    
                    # 获取目的国信息和国家代码 - 增强版：严格基于订单号对应的收件人国家
                    try:
                        # 核心匹配：严格获取订单号对应的收件人国家 - 这是唯一的主要匹配依据
                        # 优先获取Consignee Country字段，增加更多可能的字段名，提高匹配覆盖率
                        consignee_country = data_row.get("Consignee Country", "") or \
                                          data_row.get("收件人国家", "") or \
                                          data_row.get("Recipient Country", "") or \
                                          data_row.get("Country", "") or \
                                          data_row.get("国家", "") or \
                                          data_row.get("收件人国家/Consignee Country", "") or \
                                          data_row.get("收件人国家名称", "") or \
                                          data_row.get("Recipient Country Name", "") or \
                                          data_row.get("Country of Consignee", "") or \
                                          data_row.get("Consignee Country Code", "") or \
                                          data_row.get("收件人所在国家", "") or \
                                          "United States"  # 默认值作为最后的兜底
                        
                        # 获取辅助信息（仅在主匹配失败时使用）
                        destination_country = data_row.get("目的国", "") or \
                                             data_row.get("Destination Country", "") or \
                                             data_row.get("Country", "") or \
                                             data_row.get("国家", "") or ""
                        
                        country_code_field = data_row.get("国家代码", "") or \
                                           data_row.get("Country Code", "") or ""
                        
                        # 增强：添加国家代码映射表，支持更多国家
                        country_code_map = {
                            "United States": "502",
                            "USA": "502",
                            "US": "502",
                            "United Kingdom": "303",
                            "UK": "303",
                            "GB": "303",
                            "UNITED KINGDOM": "303",
                            "Germany": "304",
                            "GERMANY": "304",
                            "DE": "304",
                            "Canada": "305",
                            "CANADA": "305",
                            "CA": "305",
                            "Australia": "306",
                            "AUSTRALIA": "306",
                            "AU": "306",
                            "France": "307",
                            "FRANCE": "307",
                            "FR": "307",
                            "Italy": "308",
                            "ITALY": "308",
                            "IT": "308",
                            "Spain": "309",
                            "SPAIN": "309",
                            "ES": "309",
                            "Japan": "310",
                            "JAPAN": "310",
                            "JP": "310",
                            "South Korea": "311",
                            "KOREA, REPUBLIC OF": "311",
                            "KR": "311",
                            "Singapore": "312",
                            "SINGAPORE": "312",
                            "SG": "312",
                            "Malaysia": "313",
                            "MALAYSIA": "313",
                            "MY": "313",
                            "India": "314",
                            "INDIA": "314",
                            "IN": "314",
                            "Brazil": "315",
                            "BRAZIL": "315",
                            "BR": "315",
                            "Russia": "316",
                            "RUSSIAN FEDERATION": "316",
                            "RU": "316",
                            "South Africa": "317",
                            "SOUTH AFRICA": "317",
                            "ZA": "317"
                        }
                        
                        # 初始化country_info变量，确保在所有情况下都有定义
                        country_info = None
                        
                        # 1. 首先尝试使用国家代码映射表直接获取3字码
                        country_code = country_code_map.get(consignee_country, None)
                        is_country_valid = True  # 标记国家是否有效
                        
                        # 新增：记录匹配到的国家名称，用于后续查找申报金额规则
                        matched_country_name = "美国"  # 默认值
                        
                        if country_code:
                            # 根据国家代码反向查找国家名称
                            for name, code in country_code_map.items():
                                if code == country_code:
                                    # 使用中文国家名称，因为申报金额规则中的国家名称是中文
                                    if name in ["United States", "USA", "US"]:
                                        matched_country_name = "美国"
                                    elif name in ["United Kingdom", "UK", "GB", "UNITED KINGDOM"]:
                                        matched_country_name = "英国"
                                    elif name in ["Germany", "GERMANY", "DE"]:
                                        matched_country_name = "德国"
                                    elif name in ["Canada", "CANADA", "CA"]:
                                        matched_country_name = "加拿大"
                                    elif name in ["Australia", "AUSTRALIA", "AU"]:
                                        matched_country_name = "澳大利亚"
                                    elif name in ["France", "FRANCE", "FR"]:
                                        matched_country_name = "法国"
                                    elif name in ["Italy", "ITALY", "IT"]:
                                        matched_country_name = "意大利"
                                    elif name in ["Spain", "SPAIN", "ES"]:
                                        matched_country_name = "西班牙"
                                    elif name in ["Japan", "JAPAN", "JP"]:
                                        matched_country_name = "日本"
                                    elif name in ["South Korea", "KOREA, REPUBLIC OF", "KR"]:
                                        matched_country_name = "韩国"
                                    elif name in ["Singapore", "SINGAPORE", "SG"]:
                                        matched_country_name = "新加坡"
                                    elif name in ["Malaysia", "MALAYSIA", "MY"]:
                                        matched_country_name = "马来西亚"
                                    elif name in ["India", "INDIA", "IN"]:
                                        matched_country_name = "印度"
                                    elif name in ["Brazil", "BRAZIL", "BR"]:
                                        matched_country_name = "巴西"
                                    elif name in ["Russia", "RUSSIAN FEDERATION", "RU"]:
                                        matched_country_name = "俄罗斯"
                                    elif name in ["South Africa", "SOUTH AFRICA", "ZA"]:
                                        matched_country_name = "南非"
                                    break
                        else:
                            # 2. 尝试使用目的国进行映射
                            country_code = country_code_map.get(destination_country, None)
                            if country_code:
                                # 根据国家代码反向查找国家名称
                                for name, code in country_code_map.items():
                                    if code == country_code:
                                        if name in ["United States", "USA", "US"]:
                                            matched_country_name = "美国"
                                        elif name in ["United Kingdom", "UK", "GB", "UNITED KINGDOM"]:
                                            matched_country_name = "英国"
                                        elif name in ["Germany", "GERMANY", "DE"]:
                                            matched_country_name = "德国"
                                        break
                            else:
                                # 3. 尝试使用国家代码字段进行映射
                                country_code = country_code_map.get(country_code_field, None)
                                if country_code:
                                    # 根据国家代码反向查找国家名称
                                    for name, code in country_code_map.items():
                                        if code == country_code:
                                            if name in ["United States", "USA", "US"]:
                                                matched_country_name = "美国"
                                            elif name in ["United Kingdom", "UK", "GB", "UNITED KINGDOM"]:
                                                matched_country_name = "英国"
                                            elif name in ["Germany", "GERMANY", "DE"]:
                                                matched_country_name = "德国"
                                            break
                                else:
                                    # 4. 尝试模糊匹配国家名称
                                    matched = False
                                    for country_name, code in country_code_map.items():
                                        if country_name.lower() in consignee_country.lower() or consignee_country.lower() in country_name.lower():
                                            country_code = code
                                            matched = True
                                            # 设置匹配到的国家名称
                                            if country_name in ["United States", "USA", "US"]:
                                                matched_country_name = "美国"
                                            elif country_name in ["United Kingdom", "UK", "GB", "UNITED KINGDOM"]:
                                                matched_country_name = "英国"
                                            elif country_name in ["Germany", "GERMANY", "DE"]:
                                                matched_country_name = "德国"
                                            break
                                    
                                    if not matched:
                                        # 5. 最后尝试匹配国家代码数据中的收件人国家字段
                                        for country in self.country_codes:
                                            if isinstance(country, dict):
                                                cc_in_data = country.get("consignee_country", "").strip()
                                                if cc_in_data and (cc_in_data == consignee_country or \
                                                                 cc_in_data.lower() in consignee_country.lower() or \
                                                                 consignee_country.lower() in cc_in_data.lower()):
                                                    country_info = country
                                                    three_letter_code = country_info.get("three_letter_code")
                                                    if three_letter_code:
                                                        country_code = three_letter_code
                                                    # 使用国家数据中的国家名称
                                                    if country_info:
                                                        matched_country_name = country_info.get("country_name", "美国")
                                                    break
                                        
                                        if not country_info:
                                            # 所有匹配都失败，使用默认值
                                            is_country_valid = False
                                            matched_country_name = "美国"
                    except Exception as e:
                        # 异常处理：确保即使出错也能继续处理
                        country_code = "USA"  # 使用默认值继续
                        matched_country_name = "美国"  # 设置默认国家名称，确保后续处理正常
                    
                    # 计算申报单价
                    try:
                        # 根据国家名称查找对应的申报金额规则
                        declaration_ratio = 0.5  # 默认申报比例
                        max_declaration_amount = 800  # 默认最高申报金额
                        
                        # 使用新的matched_country_name变量，确保无论国家匹配方式如何，都能正确查找申报金额规则
                        country_name_for_declaration = matched_country_name
                        
                        for rule in self.declaration_amount_rules:
                            if isinstance(rule, dict) and rule.get("country_name") == country_name_for_declaration:
                                declaration_ratio = rule.get("declaration_ratio", 0.5)
                                max_declaration_amount = rule.get("max_declaration_amount", 800)
                                break
                        
                        # 获取订单金额，支持多种字段名，严格优先使用"总销售金额/AmountPaid"
                        # 尝试从所有可能的字段名中获取订单金额，提高匹配成功率
                        order_amount = 0.0
                        
                        # 尝试所有可能的字段名，按优先级顺序
                        possible_fields = [
                            "总销售金额/AmountPaid",
                            "销售金额/Amount",
                            "总销售金额",
                            "AmountPaid",
                            "订单金额",
                            "销售金额",
                            "总销售金额/AmountPaid  ",
                            "Total Amount",
                            "Order Amount"
                        ]
                        
                        for field in possible_fields:
                            if field in data_row:
                                # 尝试获取字段值，处理可能的空值情况
                                field_value = data_row[field]
                                if field_value is not None and field_value != "":
                                    try:
                                        order_amount = float(field_value)
                                        break
                                    except (ValueError, TypeError):
                                        continue
                        
                        # 获取包裹内单个SKC的商品数量
                        sku_quantity = company_declaration_info.get("包裹内单个SKC的商品数量", 1)
                        
                        # 确保sku_quantity是正数，避免除以零
                        if sku_quantity <= 0:
                            sku_quantity = 1
                        
                        # 计算单个商品的实际销售价格：订单总金额 / 商品数量
                        actual_unit_price = order_amount / sku_quantity
                        
                        # 计算申报单价：单个商品实际销售价格 * 申报比例
                        declared_price = actual_unit_price * declaration_ratio
                        
                        # 计算申报总价：申报单价 * 商品数量
                        declared_total_price = declared_price * sku_quantity
                        
                        # 条件判断1：如果计算后的申报单价 < 15，则使用15
                        if declared_price <= 0 or declared_price < 15.0:
                            declared_price = 15.0
                        
                        # 条件判断2：如果计算后的申报单价 > 国家对应的最高申报金额，则使用最高申报金额
                        if declared_price > max_declaration_amount:
                            declared_price = max_declaration_amount
                        
                        # 重新计算申报总价（如果申报单价被调整）
                        declared_total_price = declared_price * sku_quantity
                    except Exception as e:
                        # 计算失败，跳过该订单
                        error_orders += 1
                        # 继续处理下一个订单
                        continue
                    
                    # 获取物流跟踪号 - 增强版本：优先从物流信息中匹配
                    logistics_tracking_no = order_no  # 默认使用订单号
                    try:
                        # 首先从数据行中获取物流跟踪号
                        order_tracking_no = data_row.get("物流跟踪号", "") or \
                                           data_row.get("Tracking Number", "") or \
                                           data_row.get("包裹号（运单号）", "")
                        
                        # 如果数据行中有物流跟踪号，先使用它
                        if order_tracking_no:
                            logistics_tracking_no = order_tracking_no
                        
                        # 尝试从已匹配的物流信息中获取更准确的物流跟踪号
                        if hasattr(self, 'logistics_data') and self.logistics_data:
                            for logistics_row in self.logistics_data:
                                # 检查物流信息是否与当前订单匹配（通过订单号/平台订单号）
                                platform_order_no = logistics_row.get("平台订单号", "")
                                # 检查是否已经匹配到当前订单
                                if platform_order_no and (platform_order_no == order_no or \
                                                        ("is_matched" in logistics_row and logistics_row["is_matched"] and \
                                                         (logistics_row.get("订单号") == order_no or logistics_row.get("Order Code") == order_no))):
                                    
                                    # 从物流信息中获取跟踪号（可能有不同的字段名）
                                    logistics_info_tracking_no = logistics_row.get("物流跟踪号", "") or \
                                                              logistics_row.get("Tracking Number", "") or \
                                                              logistics_row.get("包裹号（运单号）", "") or \
                                                              logistics_row.get("运单号", "") or \
                                                              logistics_row.get("包裹号", "")
                                    
                                    if logistics_info_tracking_no:
                                        logistics_tracking_no = logistics_info_tracking_no
                                        break
                        
                        # 预处理跟踪号：去掉可能的前缀
                        if logistics_tracking_no.startswith("TRACK-"):
                            logistics_tracking_no = logistics_tracking_no[6:]
                    except Exception as e:
                        # 出错时使用订单号作为跟踪号
                        logistics_tracking_no = order_no
                    
                    # 生成报关单数据 - 严格按照README.md中的字段匹配规则
                    try:
                        # 确保company_declaration_info是字典类型
                        if not isinstance(company_declaration_info, dict):
                            company_declaration_info = {}
                        
                        # 确保所有字段都从匹配的报关信息中获取，而不是使用硬编码固定值
                        declaration_item = {
                            # 报关单各字段匹配：订单号所属店铺的公司的报关信息
                            "提单号": company_declaration_info.get("提单号", ""),
                            "订单编号": order_no,  # 订单编号直接使用原始订单数据
                            "大箱号": company_declaration_info.get("大箱号", ""),  # 允许报关信息中设置大箱号
                            "快件单号": company_declaration_info.get("快件单号", ""),  # 允许报关信息中设置快件单号
                            "包裹号（运单号）": logistics_tracking_no,  # 报关单的包裹号对应物流信息的物流跟踪号
                            "目的国": country_code,  # 报关单字段的目的国对应国家代码的3字码
                            "SKC": order_no,  # 同订单编号
                            "敏感品类别": company_declaration_info.get("敏感品类别", ""),  # 允许报关信息中设置敏感品类别
                            "商品品名": company_declaration_info.get("商品品名", ""),
                            "HS CODE": company_declaration_info.get("HS CODE", ""),
                            "规格型号": company_declaration_info.get("规格型号", ""),
                            "包裹内单个SKC的商品数量": company_declaration_info.get("包裹内单个SKC的商品数量", 1),
                            "申报单位": company_declaration_info.get("申报单位", ""),
                            "商品申报单价": declared_price,  # 根据申报金额规则计算（这是需要动态计算的字段）
                            "申报币制": company_declaration_info.get("申报币制", ""),
                            "商品总净重(KG)": company_declaration_info.get("商品总净重(KG)", 0),
                            "第一法定数量": company_declaration_info.get("第一法定数量", 0),
                            "第一法定单位": company_declaration_info.get("第一法定单位", ""),
                            "第二法定数量": company_declaration_info.get("第二法定数量", ""),
                            "第二法定单位": company_declaration_info.get("第二法定单位", ""),
                            "电商企业代码": company_declaration_info.get("电商企业代码", ""),
                            "电商企业名称": company_declaration_info.get("电商企业名称", ""),
                            "电商平台代码": company_declaration_info.get("电商平台代码", ""),
                            "电商平台名称": company_declaration_info.get("电商平台名称", ""),
                            "收款企业代码": company_declaration_info.get("收款企业代码", ""),
                            "收款企业名称": company_declaration_info.get("收款企业名称", ""),
                            "生产企业代码": company_declaration_info.get("生产企业代码", ""),
                            "生产企业名称": company_declaration_info.get("生产企业名称", ""),
                            "电商企业dxpId": company_declaration_info.get("电商企业dxpId", "")
                        }
                        
                        # 验证必填字段，确保关键信息不为空
                        required_fields = ["商品品名", "HS CODE", "申报单位", "电商企业名称"]
                        for field in required_fields:
                            if not declaration_item.get(field):
                                # 为关键字段设置最小化默认值，以保证数据有效性
                                if field == "商品品名":
                                    declaration_item[field] = "默认商品"
                                elif field == "HS CODE":
                                    declaration_item[field] = "000000"
                                elif field == "申报单位":
                                    declaration_item[field] = "个"
                                elif field == "电商企业名称":
                                    declaration_item[field] = company_name  # 使用已匹配的公司名称
                        
                        # 添加生成时间戳，方便追踪
                        declaration_item["生成时间"] = ""
                        
                        # 添加国家有效性标记，用于前端显示红色提示
                        declaration_item["is_country_valid"] = is_country_valid
                        
                        # 添加到结果列表
                        declaration_data.append(declaration_item)
                        processed_orders += 1
                    except Exception as e:
                        error_orders += 1
                        continue
                
                except Exception as e:
                    # 单个订单处理失败，记录错误并继续处理下一个订单
                    error_orders += 1
                    continue
            
            # 保存生成的报关单数据到实例变量
            self.customs_data = declaration_data
            
            return declaration_data
            
        except Exception as e:
            # 捕获整个生成过程中的异常
            # 即使发生系统性错误，也返回已生成的数据
            return declaration_data
    
    def get_order_data(self):
        """获取订单数据
        
        Returns:
            list: 订单数据列表
        """
        return self.order_data
    
    def get_declaration_template_fields(self):
        """获取报关单模板字段
        
        Returns:
            list: 报关单模板字段列表
        """
        return self.declaration_template_fields
    
    def get_declaration_info(self):
        """获取报关信息
        
        Returns:
            list: 报关信息列表
        """
        return self.declaration_info
    
    def add_declaration_info(self, declaration_data):
        """添加报关信息
        
        Args:
            declaration_data: 报关信息字典
        """
        try:
            # 添加报关信息到列表
            self.declaration_info.append(declaration_data)
            print("报关信息添加成功")
        except Exception as e:
            print(f"添加报关信息时发生错误: {e}")
            raise
    

    
    def get_customs_data(self):
        """获取报关单数据
        
        Returns:
            list: 报关单数据列表
        """
        return self.customs_data
    
    def export_customs_data(self, customs_data, file_path):
        """导出报关单数据到Excel文件
        
        Args:
            customs_data: 报关单数据列表
            file_path: 导出文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "报关单数据"
            
            if customs_data:
                # 获取表头
                headers = list(customs_data[0].keys())
                
                # 写入表头
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                
                # 写入数据
                for data in customs_data:
                    row = [data.get(header, "") for header in headers]
                    ws.append(row)
            
            # 保存文件
            wb.save(file_path)
            print(f"报关单数据导出成功: {file_path}")
        except Exception as e:
            print(f"导出报关单数据失败: {e}")
            raise
    
    def export_customs_data_by_company(self, customs_data, export_dir):
        """按公司导出报关单数据，每个公司一个Excel文件
        
        Args:
            customs_data: 报关单数据列表
            export_dir: 导出目录路径
        """
        try:
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            # 按公司分组数据
            company_data = {}
            for data in customs_data:
                company_name = data.get("电商企业名称", "未知公司")
                if company_name not in company_data:
                    company_data[company_name] = []
                company_data[company_name].append(data)
            
            # 为每个公司创建Excel文件
            for company_name, data_list in company_data.items():
                # 创建工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = f"{company_name}报关单数据"
                
                if data_list:
                    # 获取表头
                    headers = list(data_list[0].keys())
                    
                    # 写入表头
                    ws.append(headers)
                    
                    # 设置表头样式
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    
                    # 写入数据
                    for data in data_list:
                        row = [data.get(header, "") for header in headers]
                        ws.append(row)
                
                # 保存文件
                file_name = f"{company_name}_报关单数据.xlsx"
                file_path = os.path.join(export_dir, file_name)
                wb.save(file_path)
                print(f"公司 {company_name} 报关单数据导出成功: {file_path}")
        except Exception as e:
            print(f"按公司导出报关单数据失败: {e}")
            raise
    
    def add_country_code(self, consignee_country, three_letter_code):
        """添加国家代码
        
        Args:
            consignee_country: 收件人国家/Consignee Country
            three_letter_code: 3字码
        """
        try:
            # 检查是否已存在
            for code in self.country_codes:
                if code["consignee_country"] == consignee_country:
                    print(f"国家代码已存在: {consignee_country}")
                    return
            
            # 创建国家代码数据字典
            country_code_data = {
                "consignee_country": consignee_country,
                "three_letter_code": three_letter_code
            }
            
            # 添加到列表
            self.country_codes.append(country_code_data)
            print(f"国家代码添加成功: {consignee_country}")
            # 保存到文件
            self.save_country_codes()
        except Exception as e:
            print(f"添加国家代码时发生错误: {e}")
            raise
    
    def edit_country_code(self, original_consignee_country, new_consignee_country, new_three_letter_code):
        """编辑国家代码
        
        Args:
            original_consignee_country: 原始收件人国家
            new_consignee_country: 新收件人国家/Consignee Country
            new_three_letter_code: 新3字码
        """
        try:
            # 查找要编辑的国家代码
            for i, country_code in enumerate(self.country_codes):
                if country_code["consignee_country"] == original_consignee_country:
                    # 更新国家代码
                    self.country_codes[i] = {
                        "consignee_country": new_consignee_country,
                        "three_letter_code": new_three_letter_code
                    }
                    print("国家代码编辑成功")
                    # 保存到文件
                    self.save_country_codes()
                    return
            
            # 如果没有找到匹配的国家代码
            print(f"未找到国家代码: {original_consignee_country}")
        except Exception as e:
            print(f"编辑国家代码时发生错误: {e}")
            raise
    
    def get_currency_data(self):
        """获取币种数据
        
        Returns:
            list: 币种数据列表
        """
        return self.currency_data
    
    def get_shop_company_data(self):
        """获取店铺对应公司数据
        
        Returns:
            list: 店铺对应公司数据列表
        """
        return self.shop_company_data
    
    def add_shop_company(self, shop_name, company_name):
        """添加店铺对应公司
        
        Args:
            shop_name: 店铺名称
            company_name: 所属公司
        """
        try:
            # 创建店铺对应公司数据字典
            shop_company_data = {
                "shop_name": shop_name,
                "company_name": company_name
            }
            
            # 添加到列表
            self.shop_company_data.append(shop_company_data)
            print("店铺对应公司添加成功")
            # 保存到文件
            self.save_shop_company_data()
        except Exception as e:
            print(f"添加店铺对应公司时发生错误: {e}")
            raise
    
    def get_country_codes(self):
        """获取国家代码
        
        Returns:
            list: 国家代码列表
        """
        return self.country_codes
    
    def update_country_code(self, old_consignee_country, new_consignee_country):
        """更新国家代码
        
        Args:
            old_consignee_country: 旧收件人国家
            new_consignee_country: 新收件人国家
        """
        try:
            # 查找并更新
            for code in self.country_codes:
                if code["consignee_country"] == old_consignee_country:
                    code["consignee_country"] = new_consignee_country
                    print(f"国家代码更新成功: {old_consignee_country} -> {new_consignee_country}")
                    # 保存到文件
                    self.save_country_codes()
                    return
            print(f"未找到国家代码: {old_consignee_country}")
        except Exception as e:
            print(f"更新国家代码时发生错误: {e}")
            raise
    
    def delete_country_code(self, consignee_country):
        """删除国家代码
        
        Args:
            consignee_country: 收件人国家
        """
        try:
            # 查找并删除
            for i, code in enumerate(self.country_codes):
                if code.get("consignee_country") == consignee_country:
                    del self.country_codes[i]
                    print(f"国家代码删除成功: {consignee_country}")
                    # 保存到文件
                    self.save_country_codes()
                    return
            print(f"未找到国家代码: {consignee_country}")
        except Exception as e:
            print(f"删除国家代码时发生错误: {e}")
            raise
    
    def save_country_codes(self):
        """保存国家代码到文件"""
        try:
            with open("country_codes.json", "w", encoding="utf-8") as f:
                json.dump(self.country_codes, f, ensure_ascii=False, indent=4)
            print("国家代码保存成功")
        except Exception as e:
            print(f"保存国家代码时发生错误: {e}")
    
    def load_country_codes(self):
        """从文件加载国家代码"""
        try:
            if os.path.exists("country_codes.json"):
                with open("country_codes.json", "r", encoding="utf-8") as f:
                    self.country_codes = json.load(f)
                print("国家代码加载成功")
        except Exception as e:
            print(f"加载国家代码时发生错误，使用默认值: {e}")
    
    def save_currency_data(self):
        """保存币种数据到文件"""
        try:
            with open("currency_data.json", "w", encoding="utf-8") as f:
                json.dump(self.currency_data, f, ensure_ascii=False, indent=4)
            print("币种数据保存成功")
        except Exception as e:
            print(f"保存币种数据时发生错误: {e}")
    
    def load_currency_data(self):
        """从文件加载币种数据"""
        try:
            if os.path.exists("currency_data.json"):
                with open("currency_data.json", "r", encoding="utf-8") as f:
                    self.currency_data = json.load(f)
                print("币种数据加载成功")
        except Exception as e:
            print(f"加载币种数据时发生错误，使用默认值: {e}")
    
    def save_shop_company_data(self):
        """保存店铺对应公司数据到文件"""
        try:
            with open("shop_company_data.json", "w", encoding="utf-8") as f:
                json.dump(self.shop_company_data, f, ensure_ascii=False, indent=4)
            print("店铺对应公司数据保存成功")
        except Exception as e:
            print(f"保存店铺对应公司数据时发生错误: {e}")
    
    def load_shop_company_data(self):
        """从文件加载店铺对应公司数据"""
        try:
            if os.path.exists("shop_company_data.json"):
                with open("shop_company_data.json", "r", encoding="utf-8") as f:
                    self.shop_company_data = json.load(f)
                print("店铺对应公司数据加载成功")
        except Exception as e:
            print(f"加载店铺对应公司数据时发生错误，使用默认值: {e}")
    
    def get_declaration_amount_rules(self):
        """获取申报金额规则
        
        Returns:
            list: 申报金额规则列表
        """
        return self.declaration_amount_rules
    
    def add_declaration_amount_rule(self, country_name, declaration_ratio, max_declaration_amount):
        """添加申报金额规则
        
        Args:
            country_name: 国家名称
            declaration_ratio: 申报比例
            max_declaration_amount: 最高申报金额
        """
        try:
            # 检查是否已存在
            for rule in self.declaration_amount_rules:
                if rule["country_name"] == country_name:
                    print(f"申报金额规则已存在: {country_name}")
                    return
            
            # 创建申报金额规则数据字典
            rule_data = {
                "country_name": country_name,
                "declaration_ratio": declaration_ratio,
                "max_declaration_amount": max_declaration_amount
            }
            
            # 添加到列表
            self.declaration_amount_rules.append(rule_data)
            print(f"申报金额规则添加成功: {country_name} - 比例: {declaration_ratio}, 最高金额: {max_declaration_amount}")
            # 保存到文件
            self.save_declaration_amount_rules()
        except Exception as e:
            print(f"添加申报金额规则时发生错误: {e}")
            raise
    
    def update_declaration_amount_rule(self, old_country_name, new_country_name, declaration_ratio, max_declaration_amount):
        """更新申报金额规则
        
        Args:
            old_country_name: 旧国家名称
            new_country_name: 新国家名称
            declaration_ratio: 申报比例
            max_declaration_amount: 最高申报金额
        """
        try:
            # 查找并更新
            for rule in self.declaration_amount_rules:
                if rule["country_name"] == old_country_name:
                    rule["country_name"] = new_country_name
                    rule["declaration_ratio"] = declaration_ratio
                    rule["max_declaration_amount"] = max_declaration_amount
                    print(f"申报金额规则更新成功: {old_country_name} -> {new_country_name} - 比例: {declaration_ratio}, 最高金额: {max_declaration_amount}")
                    # 保存到文件
                    self.save_declaration_amount_rules()
                    return
            print(f"未找到申报金额规则: {old_country_name}")
        except Exception as e:
            print(f"更新申报金额规则时发生错误: {e}")
            raise
    
    def delete_declaration_amount_rule(self, country_name):
        """删除申报金额规则
        
        Args:
            country_name: 国家名称
        """
        try:
            # 查找并删除
            for i, rule in enumerate(self.declaration_amount_rules):
                if rule["country_name"] == country_name:
                    del self.declaration_amount_rules[i]
                    print(f"申报金额规则删除成功: {country_name}")
                    # 保存到文件
                    self.save_declaration_amount_rules()
                    return
            print(f"未找到申报金额规则: {country_name}")
        except Exception as e:
            print(f"删除申报金额规则时发生错误: {e}")
            raise
    
    def save_declaration_amount_rules(self):
        """保存申报金额规则到文件"""
        try:
            with open("declaration_amount_rules.json", "w", encoding="utf-8") as f:
                json.dump(self.declaration_amount_rules, f, ensure_ascii=False, indent=4)
            print("申报金额规则保存成功")
        except Exception as e:
            print(f"保存申报金额规则时发生错误: {e}")
    
    def load_declaration_amount_rules(self):
        """从文件加载申报金额规则"""
        try:
            if os.path.exists("declaration_amount_rules.json"):
                with open("declaration_amount_rules.json", "r", encoding="utf-8") as f:
                    self.declaration_amount_rules = json.load(f)
                print("申报金额规则加载成功")
        except Exception as e:
            print(f"加载申报金额规则时发生错误，使用默认值: {e}")
    
    def clear_sheet_data(self, sheet):
        """清空工作表数据（保留表头）
        
        Args:
            sheet: 要清空数据的工作表对象
        """
        # 获取最后一行
        max_row = sheet.max_row
        if max_row > 1:
            # 清空从第2行开始的数据
            sheet.delete_rows(2, max_row - 1)
    
    def save_template_files(self):
        """保存各个模板为新的Excel文件"""
        if self.this_workbook is None:
            return
        
        try:
            # 获取当前工作簿路径
            path = os.path.dirname(os.path.abspath(__file__))
            
            # 创建导入模板目录（如果不存在）
            import_template_dir = os.path.join(path, "导入模板")
            if not os.path.exists(import_template_dir):
                os.makedirs(import_template_dir)
            
            # 获取当前日期
            today = datetime.now().strftime("%Y-%m-%d")
            
            # 新的模板字段列表
            new_template_fields = [
                "提单号", "订单编号", "大箱号", "快件单号", "包裹号（运单号）", "目的国", 
                "SKC", "敏感品类别", "商品品名", "HS CODE", "规格型号", 
                "包裹内单个SKC的商品数量", "申报单位", "商品申报单价", "申报币制", 
                "商品总净重(KG)", "第一法定数量", "第一法定单位", "第二法定数量", 
                "第二法定单位", "电商企业代码", "电商企业名称", "电商平台代码", 
                "电商平台名称", "收款企业代码", "收款企业名称", "生产企业代码", 
                "生产企业名称", "电商企业dxpId"
            ]
            
            # 保存各个模板
            for template_name, file_name in self.TEMPLATE_FILE_MAP.items():
                # 检查模板是否有数据
                template_sheet = self.this_workbook[template_name]
                if template_sheet.max_row <= 1:
                    continue  # 没有数据，跳过
                
                # 创建新工作簿
                new_wb = Workbook()
                new_wb.remove(new_wb.active)  # 删除默认工作表
                
                # 复制模板工作表到新工作簿
                new_sheet = new_wb.create_sheet(title="报关模板统一manifest")
                
                # 添加新的模板字段作为表头
                new_sheet.append(new_template_fields)
                
                # 复制数据（从第二行开始，跳过原表头）
                for i, row in enumerate(template_sheet.iter_rows()):
                    if i == 0:  # 跳过原表头
                        continue
                    new_sheet.append([cell.value for cell in row])
                
                # 创建毛重工作表
                weight_sheet = new_wb.create_sheet(title="毛重")
                
                # 复制AD:AE列数据到毛重工作表
                for row in template_sheet.iter_rows(min_col=30, max_col=31):
                    weight_sheet.append([cell.value for cell in row])
                
                # 生成文件名
                base_filename = f"{file_name}{today}x.xlsx"
                YDPath = os.path.join(import_template_dir, base_filename)
                
                # 检查文件是否存在，如果存在则增加数字
                x = 1
                while os.path.exists(YDPath):
                    x += 1
                    # 重新生成文件名
                    YDPath = os.path.join(import_template_dir, f"{file_name}{today}x{x}.xlsx")
                
                # 保存新文件
                new_wb.save(YDPath)
                print(f"保存文件: {YDPath}")
        except Exception as e:
            print(f"保存模板文件时发生错误: {e}")

# 测试代码
if __name__ == "__main__":
    processor = ExcelProcessor()
    # 可以在这里测试各个方法
    # processor.import_logistics_file()
    # processor.process_logistics_data()
    # print(processor.get_logistics_data())
